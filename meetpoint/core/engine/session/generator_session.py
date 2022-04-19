import xlrd
from openpyxl import load_workbook
from datetime import date, time, datetime
from hashlib import sha1
from os import path


def get_session(name_input_file, input_file):
    filename = str(sha1(input_file).hexdigest())
    file_extension = name_input_file.split('.')[1]

    #if name_input_file.split('.')[-1] == 'xls':
    #    filename += '.xls'

    #elif name_input_file.split('.')[-1] == 'xlsx':
    #    filename += '.xlsx'
    
    get_data_from_file(filename, file_extension, input_file)
    
    return filename


def get_data_from_file(filename, file_extension, input_file):
    
    if not path.exists(f'sessions/{filename}'):
        with open(f'sessions/{filename}', 'wb') as file:
            if file_extension == 'xls':
                file.write(parse_xls(filename))

            elif file_extension == 'xlsx':
                file.write(parse_xlsx(filename))

            elif file_extension == 'txt' or file_extension == 'csv':
                file.write(parse_csv_txt(filename)) 
        
    
def parse_xls(filename):
    request_dict['endpoints_and_quantity_people'] = []
    book = xlrd.open_workbook(f'sessions/{filename}.xls')
    quant_sheets = book.nsheets
    
    if isinstance(cell_value(rowx=2, colx=1), int):        
        for rx in range(2, sheet.nrows):
            request_dict['endpoints_and_quantity_people'].append([cell_value(rowx=rx, colx=0), cell_value(rowx=rx, colx=1)])
    else:
        for rx in range(2, sheet.nrow):
            quantity_people = cell_value(rowx=rowx, colx=1)
            request_dict['endpoints_and_quantuti_people'].append([cell_value(rowx=rx, colx=0), quantity_people])
    
    request_dict['start_date'] = cell_value(rowx=0, colx=0).split('-')[0]

    request_dict['end_date'] = cell_value(rowx=0, colx=0).split('-')[1]
    
    return request_dict


def parse_xlsx(filename):
    request_dict = {}
    request_dict['endpoints_and_quantity_people'] = []
    wb = load_workbook(f'sessions/{filename}.xlsx').active

    if isinstance(wb.cell(row=3, column=2).value, int):
        for row in range(3, wb.max_row+1):
            request_dict['endpoints_and_quantity_people'].append([wb.cell(row=row, column=1).value, wb.cell(row=row, column=2).value])
    else:
        for row in range(3, wb.max_row+1):
            quantity_people = len(wb.cell(row=row, column=2).value.split(','))
            request_dict['endpoints_and_quantity_people'].append([wb.cell(row=row, column=1).value, quantity_people])

    request_dict['start_date'] = wb.cell(row=1, column=1).value.split('-')[0]
    request_dict['end_date'] = wb.cell(row=1, column=1).value.split('-')[1]
    
    return request_dict
    
    
def parse_csv_txt(filename):
    request_dict = {}
    request_dict['endpoints_and_quantity_people'] = []

    with open(f'sessions/{filename}', 'r') as file:
        content = str(file.read())
        if content[-1] == '\n':
            content = content[:-1]

        content = content.split('\n')

    if content[1][-1].isdigit():
        for row in range(1, len(content)):
            content[row] = content[row].split(', ')
            request_dict['endpoints_and_quantity_people'].append(content[row])
    else:
        for row in range(1, len(content)):
            quantity_people = len(content[row].split(', '))-1
            content[row] = [content[row].split(', ')[0], quantity_people]
            request_dict['endpoints_and_quantity_people'].append([content[row][0], quantity_people])
    
    request_dict['start_date'] = content[0].split('-')[0]
    request_dict['end_date'] = content[0].split('-')[1]
    
    return request_dict
