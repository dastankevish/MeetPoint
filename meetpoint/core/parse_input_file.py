import xlrd
from openpyxl import load_workbook
from datetime import date, time, datetime
from hashlib import sha1
from os import path


def get_data_from_file(name_input_file, input_file):
    filename = str(sha1(input_file).hexdigest())

    if name_input_file.split('.')[-1] == 'xls':
        filename += '.xls'

    elif name_input_file.split('.')[-1] == 'xlsx':
        filename += '.xlsx'


    if not path.exists(f'sessions/{filename}'):
        with open(f'sessions/{filename}', 'wb') as file:
            file.write(input_file)

    
    if name_input_file.split('.')[-1] =='xls':
        return parse_xls(filename)

    elif name_input_file.split('.')[-1] =='xlsx':
        return parse_xlsx(filename)

    elif name_input_file.split('.')[-1] == 'txt' or name_input_file.split('.')[-1] == 'csv':
        return parse_csv_txt(filename)
    
    elif name_input_file.split('.')[-1] == 'json':
        return parse_json(filename)
    
    
def parse_xls(filename):
    request_dict['endpoints_and_quantity_people'] = []
    book = xlrd.open_workbook(f'sessions/{filename}')
    quant_sheets = book.nsheets
    
    #for index in range(book.nsheets):
        #sheet = book.sheet_by_index(index)
    
    if isinstance(cell_value(rowx=1, colx=1), int):        
        for rx in range(1, sheet.nrows):
            request_dict['endpoints_and_quantity_people'].append([cell_value(rowx=rx, colx=0), cell_value(rowx=rx, colx=1)])
    else:
        for rx in range(1, sheet.nrow):
            quantity_people = cell_value(rowx=rowx, colx=1)
            request_dict['endpoints_and_quantuti_people'].append([cell_value(rowx=rx, colx=0), quantity_people])
    
    request_dict['start_date'] = cell_value(rowx=1, colx=2)
    request_dict['end_date'] = cell_value(rowx=2, colx=3)
    
    return request_dict


def parse_xlsx(filename):
    request_dict = {}
    request_dict['endpoints_and_quantity_people'] = []
    wb = load_workbook(f'sessions/{filename}').active

    if isinstance(wb.cell(row=2, column=2).value, int):
        for row in range(2, wb.max_row):
            request_dict['endpoints_and_quantity_people'].append([wb.cell(row=row, column=1).value, wb.cell(row=row, column=2).value])
    else:
        for row in range(2, wb.max_row):
            quantity_people = len(wb.cell(row=row, column=2).value.split(','))
            request_dict['endpoints_and_quantity_people'].append([wb.cell(row=row, column=1).value, quantity_people])

    request_dict['start_date'] = wb.cell(row=2, column=3).value
    request_dict['end_date'] = wb.cell(row=2, column=4).value
    
    return request_dict
    
    
def parse_csv_txt(filename):
    request_dict = {}
    request_dict['endpoints_and_quantity_people'] = []

    with open(f'sessions/{filename}', 'r') as file:
        content = str(file.read()).split('\n')

    if content[0][-1].isdigit():
        for row in range(len(content)-3):
            request_dict['endpoints_and_quantity_people'].append(content[row].split(', '))
    else:
        for row in range(len(content)-3):
            quantity_people = len(content[row].split(', '))-1
            request_dict['endpoints_and_quantity_people'].append([content[row].split(', ')[0], quantity_people])
    
    request_dict['start_date'] = content[-3]
    request_dict['end_date'] = content[-2]
    
    return request_dict


def parse_json(filename):
    request_dict = json.load(f'sessions/{filename}')
    return request_dict
